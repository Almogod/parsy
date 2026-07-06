"""
Parsy Backend — Structured Data Parser
Handles CSV, JSON, XML, XLSX with schema detection and normalization.
"""
import io, csv, json, re, logging
from base_parser import ParsedBlock, FastParseResult
from base_parser import BaseParser, CorruptFileError
import chardet

log = logging.getLogger("parsy.parsers.structured")


class StructuredParser(BaseParser):

    supported_extensions = frozenset({
        "csv", "json", "xml", "svg", "xlsx", "xls", "ods",
    })

    def __init__(self) -> None:
        super().__init__()

    # Called by BaseParser.parse — do NOT call directly.
    async def _parse(self, filename: str, data: bytes) -> FastParseResult:
        ext = self.extension_of(filename)
        try:
            if ext == "csv":
                return self._parse_csv(data)
            elif ext == "json":
                return self._parse_json(data)
            elif ext in ("xml", "svg"):
                return self._parse_xml(data)
            elif ext in ("xlsx", "xls", "ods"):
                return self._parse_excel(data, ext)
            # Fallback: decode as UTF-8 text
            log.warning(
                "StructuredParser received unexpected extension; decoding as UTF-8",
                extra={"file_name": filename, "ext": ext},
            )
            return FastParseResult([], 1, data.decode("utf-8", "replace"), [], {})
        except Exception as exc:
            raise CorruptFileError(
                f"Structured parse failed: {exc}", filename=filename, cause=exc
            ) from exc

    def _parse_csv(self, data: bytes) -> FastParseResult:
        enc = chardet.detect(data)["encoding"] or "utf-8"
        text = data.decode(enc, errors="replace")

        # Sniff delimiter
        try:
            dialect = csv.Sniffer().sniff(text[:2048])
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text), dialect)
        rows = [row for row in reader if any(c.strip() for c in row)]
        if not rows:
            return FastParseResult([], 1, text, [], {})

        headers = rows[0]
        schema  = self._infer_schema(rows[1:], headers)
        blocks  = [
            ParsedBlock("heading", 1, "CSV Data", 0),
            ParsedBlock("paragraph", 0, f"{len(rows)-1} records · {len(headers)} columns", 0),
            ParsedBlock("paragraph", 0, f"Schema: {schema}", 0),
            ParsedBlock("table", 0, "", 0, table_data=rows),
        ]
        meta = {
            "rows": len(rows) - 1,
            "columns": len(headers),
            "headers": ", ".join(headers),
            "schema": schema,
        }
        return FastParseResult(blocks, 1, text, [rows], meta)

    def _infer_schema(self, rows: list, headers: list) -> str:
        if not rows: return "unknown"
        types = []
        for ci, h in enumerate(headers):
            vals = [r[ci] for r in rows if ci < len(r) and r[ci].strip()]
            if not vals:
                types.append(f"{h}:null")
                continue
            if all(re.match(r"^-?\d+$", v) for v in vals[:20]):
                types.append(f"{h}:int")
            elif all(re.match(r"^-?\d*\.\d+$", v) for v in vals[:20]):
                types.append(f"{h}:float")
            elif all(re.match(r"\d{4}-\d{2}-\d{2}", v) for v in vals[:20]):
                types.append(f"{h}:date")
            else:
                types.append(f"{h}:str")
        return "{" + ", ".join(types) + "}"

    def _parse_json(self, data: bytes) -> FastParseResult:
        try:
            obj = json.loads(data)
        except json.JSONDecodeError as e:
            return FastParseResult(
                [ParsedBlock("paragraph", 0, f"JSON Parse Error: {e}", 0)],
                1, data.decode("utf-8","replace"), [], {"error": str(e)}
            )

        pretty  = json.dumps(obj, indent=2, ensure_ascii=False)
        schema  = self._json_schema(obj)
        tables  = self._json_to_tables(obj)
        blocks  = [
            ParsedBlock("heading", 1, "JSON Data", 0),
            ParsedBlock("paragraph", 0, f"Root type: {type(obj).__name__}", 0),
            ParsedBlock("paragraph", 0, f"Schema preview: {schema}", 0),
        ]
        if tables:
            for tbl in tables:
                blocks.append(ParsedBlock("table", 0, "", 0, table_data=tbl))

        meta = {
            "rootType": type(obj).__name__,
            "keys": ", ".join(obj.keys()) if isinstance(obj, dict) else f"{len(obj)} items",
            "schema": schema,
        }
        return FastParseResult(blocks, 1, pretty, tables, meta)

    def _json_schema(self, obj, depth=0) -> str:
        if depth > 3: return "..."
        if isinstance(obj, dict):
            inner = ", ".join(f"{k}: {self._json_schema(v, depth+1)}"
                              for k, v in list(obj.items())[:5])
            return "{" + inner + ("..." if len(obj) > 5 else "") + "}"
        elif isinstance(obj, list):
            item_type = self._json_schema(obj[0], depth+1) if obj else "any"
            return f"[{item_type}]×{len(obj)}"
        return type(obj).__name__

    def _json_to_tables(self, obj) -> list:
        tables = []
        if isinstance(obj, list) and obj and isinstance(obj[0], dict):
            headers = list(obj[0].keys())
            rows    = [headers] + [[str(row.get(h,"")) for h in headers] for row in obj[:200]]
            tables.append(rows)
        return tables

    def _parse_xml(self, data: bytes) -> FastParseResult:
        from lxml import etree
        try:
            root = etree.fromstring(data)
        except Exception as e:
            return FastParseResult(
                [ParsedBlock("paragraph",0,f"XML Error: {e}",0)], 1, "", [], {}
            )
        lines = []
        self._walk_xml(root, lines, 0)
        text = "\n".join(lines)
        return FastParseResult(
            [ParsedBlock("paragraph",0,l,0) for l in lines],
            1, text, [], {"rootTag": root.tag, "childCount": len(root)}
        )

    def _walk_xml(self, el, lines, depth):
        indent  = "  " * depth
        tag     = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        text_val = (el.text or "").strip()
        if text_val:
            lines.append(f"{indent}{tag}: {text_val}")
        else:
            lines.append(f"{indent}<{tag}>")
        for child in el:
            self._walk_xml(child, lines, depth + 1)

    def _parse_excel(self, data: bytes, ext: str) -> FastParseResult:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        all_tables = []
        blocks     = []

        for sheet_name in wb.sheetnames:
            ws   = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                str_row = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in str_row):
                    rows.append(str_row)
            if rows:
                all_tables.append(rows)
                blocks.append(ParsedBlock("heading", 2, f"Sheet: {sheet_name}", 0))
                blocks.append(ParsedBlock("table", 0, "", 0, table_data=rows))

        meta = {"sheets": ", ".join(wb.sheetnames), "sheetCount": len(wb.sheetnames)}
        raw  = "\n".join(r for b in blocks if b.block_type == "heading" for r in [b.content])
        return FastParseResult(blocks, 1, raw, all_tables, meta)
